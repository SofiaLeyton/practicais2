from django.http import JsonResponse, HttpResponse
from pymongo import MongoClient
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
import io
import openpyxl
import mysql.connector
import textwrap
import jwt
from django.http import JsonResponse
from django.conf import settings
from microservicio_reportes.utils.decorators import admin_required

# Conexión a MongoDB
client = MongoClient("mongodb://localhost:27017/")
db = client['microservicios']
productos_collection = db['productos']
pedidos_collection = db['pedidos']

# 🔹 Conexión a MySQL
mysql_conn = mysql.connector.connect(
    host="127.0.0.1",
    port=3306,
    user="root",
    password="",  # vacío como en tu .env
    database="laravel"
)

def verificar_token(request):
    auth_header = request.headers.get("Authorization", None)
    if not auth_header:
        return JsonResponse({"error": "Token no proporcionado"}, status=401)

    try:
        token = auth_header.split(" ")[1]  # "Bearer <token>"
        payload = jwt.decode(token, settings.JWT_SECRET, algorithms=["HS256"])
        
        if payload.get("role") != "admin":
            return JsonResponse({"error": "Acceso denegado, no eres administrador"}, status=403)

        # Si pasa la verificación
        return JsonResponse({"mensaje": "Token válido", "usuario": payload["email"]})

    except jwt.ExpiredSignatureError:
        return JsonResponse({"error": "Token expirado"}, status=401)
    except jwt.InvalidTokenError:
        return JsonResponse({"error": "Token inválido"}, status=401)

#el admin_required va a verificar que el usuario sea admin antes de que pueda generar cualquier reporte
@admin_required
def reporte_productos_pdf(request):
    productos = list(db["productos"].find())
    buffer = io.BytesIO()
    pdf = canvas.Canvas(buffer, pagesize=letter)
    pdf.setTitle("Reporte de Productos")

    pdf.drawString(100, 750, "📦 Reporte de Productos")
    y = 720
    for p in productos:
        texto = f"Nombre: {p.get('nombre', '')} | Precio: {p.get('precio', '')} | Cantidad: {p.get('cantidad', '')}"
        pdf.drawString(80, y, texto)
        y -= 20
        if y < 50:
            pdf.showPage()
            y = 750

    pdf.save()
    buffer.seek(0)

    response = HttpResponse(buffer, content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="reporte_productos.pdf"'
    return response

@admin_required
def reporte_productos_excel(request):
    productos = list(productos_collection.find({}, {"_id": 0}))
    #crea libro de excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Productos"
    ws.append(["Nombre", "Precio", "Cantidad"])
    for producto in productos:
        ws.append([
            producto.get("nombre", "N/A"),
            producto.get("precio", 0),
            producto.get("cantidad", 0),
        ])
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="reporte_productos.xlsx"'
    return response

@admin_required
def reporte_pedidos_pdf(request):
    pedidos = list(pedidos_collection.find({}, {"_id": 0}))

    buffer = io.BytesIO()
    pdf = canvas.Canvas(buffer, pagesize=letter)
    pdf.setTitle("Reporte de Pedidos")

    pdf.setFont("Helvetica", 10)
    pdf.drawString(80, 750, "📦 Reporte de Pedidos")
    y = 730

    for pedido in pedidos:
        texto = (
            f"Cliente: {pedido.get('cliente', 'N/A')} | "
            f"Nombre producto: {pedido.get('nombre_producto', 'N/A')} | "
            f"Precio unitario: {pedido.get('precio_unitario', 0)} | "
            f"Cantidad pedida: {pedido.get('cantidad_pedida', 0)} | "
            f"Total: {pedido.get('total', 0)}"
        )

        lineas = textwrap.wrap(texto, width=100)

        for linea in lineas:
            pdf.drawString(80, y, linea)
            y -= 15
            if y < 50:
                pdf.showPage()
                pdf.setFont("Helvetica", 10)
                y = 750

        y -= 5  

    pdf.save()
    buffer.seek(0)

    response = HttpResponse(buffer, content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="reporte_pedidos.pdf"'
    return response

@admin_required
def reporte_pedidos_excel(request):
    pedidos = list(db["pedidos"].find())

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pedidos"

    ws.append(["ID", "Cliente", "Producto", "Cantidad", "Fecha"])

    for p in pedidos:
        ws.append([
            str(p.get("_id", "")),
            p.get("nombre_producto", ""),
            p.get("precio_unitario", ""),
            p.get("cantidad_pedida", ""),
            str(p.get("total", "")),
        ])

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="reporte_pedidos.xlsx"'
    return response

#este es un poco diferente, este no usa admin_required porque la autenticacion esta por dentro de la funcion
def reporte_usuarios_pdf(request):
    auth_header = request.headers.get("Authorization", None)
    token = None

    if auth_header and " " in auth_header:
        # Caso: viene en header → "Bearer <token>"
        token = auth_header.split(" ")[1]
    elif "token" in request.GET:
        # Caso: viene en la URL → ?token=<token>
        token = request.GET["token"]

    if not token:
        return JsonResponse({"error": "Token no proporcionado"}, status=401)

    try:
        # ✅ 2. Decodificar token JWT
        payload = jwt.decode(token, settings.JWT_SECRET, algorithms=["HS256"])

        # ✅ 3. Verificar rol del usuario
        if payload.get("role") != "admin":
            return JsonResponse({"error": "Acceso denegado. Solo administradores."}, status=403)

    except jwt.ExpiredSignatureError:
        return JsonResponse({"error": "Token expirado"}, status=401)
    except jwt.InvalidTokenError:
        return JsonResponse({"error": "Token inválido"}, status=401)
    except Exception as e:
        return JsonResponse({"error": f"Error al procesar token: {str(e)}"}, status=400)

    # ✅ 4. Si el token es válido, generar el PDF
    try:
        cursor = mysql_conn.cursor(dictionary=True)
        cursor.execute("SELECT id, name, email, created_at FROM users")
        usuarios = cursor.fetchall()
        cursor.close()
    except Exception as e:
        return JsonResponse({"error": f"Error al consultar base de datos: {str(e)}"}, status=500)

    # ✅ 5. Crear PDF en memoria
    buffer = io.BytesIO()
    pdf = canvas.Canvas(buffer, pagesize=letter)
    pdf.setTitle("Reporte de Usuarios")

    pdf.drawString(100, 750, "👤 Reporte de Usuarios del Sistema")
    y = 720
    for usuario in usuarios:
        texto = f"ID: {usuario['id']} | Nombre: {usuario['name']} | Email: {usuario['email']}"
        pdf.drawString(80, y, texto)
        y -= 20
        if y < 50:  # Si se acaba la página, crear una nueva
            pdf.showPage()
            y = 750

    pdf.save()
    buffer.seek(0)

    # ✅ 6. Devolver PDF como respuesta
    response = HttpResponse(buffer, content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="reporte_usuarios.pdf"'
    return response

@admin_required
def reporte_usuarios_excel(request):
    """Genera un archivo Excel con la lista de usuarios."""
    cursor = mysql_conn.cursor(dictionary=True)
    cursor.execute("SELECT id, name, email, created_at FROM users")
    usuarios = cursor.fetchall()
    cursor.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Usuarios"

    # Encabezados
    ws.append(["ID", "Nombre", "Email", "Fecha de creación"])

    # Datos
    for u in usuarios:
        ws.append([u["id"], u["name"], u["email"], str(u["created_at"])])

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename=\"reporte_usuarios.xlsx\"'
    return response